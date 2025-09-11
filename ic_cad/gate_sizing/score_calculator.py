#!/usr/bin/env python3
# =========================================================================
# PPA評分計算器 - 計算曼哈頓距離(D)、線長(WL)與最終得分(S)
# 完全依照比賽評分公式: S = (1000 × P) − (50 × D) − (300 × R)
# =========================================================================

import re, os
import argparse, datetime

def get_openroad_report(report_script, env_vars):
    """執行 OpenROAD 並擷取 TNS/WNS/Power 資訊"""
    import subprocess

    def run_single_analysis(prefix, def_file):
        env = os.environ.copy()  
        env.update(env_vars)  
        env['PREFIX'] = prefix  
        env['DEF'] = def_file  

        try:
            print("[INFO] 執行 OpenROAD 報告腳本:", report_script)
            result = subprocess.run(
                ["openroad", "-exit", "-no_init", report_script],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                universal_newlines=True,
                check=True,
                env=env
            )
            return result.stdout  
        except subprocess.CalledProcessError as e:  
            print(f"[ERROR] {prefix} OpenROAD 執行失敗")  
            print(e.stdout)  
            print(e.stderr)  
            raise  

    # 分別執行兩個設計  
    seed_output = run_single_analysis("SEED", env_vars.get("SEED_DEF"))  
    sol_output  = run_single_analysis("SOL", env_vars.get("SOL_DEF"))  

    def extract(output, tag):  
        match = re.search(f"{tag}:\\s*(-?\\d+(?:\\.\\d+)?(?:[eE][+-]?\\d+)?)", output) 
        return float(match.group(1)) if match else 0.0  
      
    report = {  
        "tns_init":    extract(seed_output, "SEED_TNS"),  
        "wns_init":    extract(seed_output, "SEED_WNS"),   
        "power_init":  extract(seed_output, "SEED_Total_Power"),
        "area_init":   extract(seed_output, "SEED_AREA"), 
        "util_init":   extract(seed_output, "SEED_UTIL"), 
        "tns_final":   extract(sol_output,  "SOL_TNS"),  
        "wns_final":   extract(sol_output,  "SOL_WNS"),  
        "power_final": extract(sol_output,  "SOL_Total_Power"),
        "area_final":  extract(sol_output,  "SOL_AREA"),
        "util_final":  extract(sol_output,  "SOL_UTIL")
    }  
      
    print("[INFO] OpenROAD report 成功解讀:", report, "\n")  
    return report

def parse_def(def_file):
    """解析DEF檔案，提取單元位置和網路資訊"""
    cells = {}  # 儲存單元位置 {cell_name: (x, y)}
    nets = []   # 儲存網路資訊 [(net_name, [pin1, pin2, ...])]
    source_timing_cells = set() ### 如果不需要可以刪

    state = None          # None / COMP / NET
    current_net   = None  # 目前正在收集的 net 名稱
    current_pins  = []    # 目前 net 的 (cell, pin) 清單

    print(f"[INFO] 解析DEF檔案: {def_file}")

    with open(def_file, 'r') as f:
        for raw in f:
            line = raw.strip()
            # ---------- 區段切換 ----------
            if line.startswith('COMPONENTS'):
                state = 'COMP'
                continue
            elif line.startswith('END COMPONENTS'):
                state = None
                continue
            elif line.startswith('NETS'):
                state = 'NET'
                continue
            elif line.startswith('END NETS'):
                state = None
                if current_net:
                    nets.append((current_net, current_pins))
                    current_net = None
                    current_pins = []
                continue
            # ---------- COMPONENTS ----------
            if state == 'COMP':
                if line.startswith('-'):
                    # 移除 HALO（若存在）
                    clean_line = line.split('+ HALO')[0].strip()
                    # 只抓以  - cellName libCellName + PLACED ( x y ) N ; 開頭的行
                    match = re.match(r'-\s+(\S+)\s+\S+(.*?)(?:\+ PLACED\s*\(\s*(-?\d+)\s+(-?\d+)\s*\)\s+\S+)?\s*;?$', clean_line)
                    if match:
                        cell_name, props, x, y = match.groups()
                        x = int(x) if x is not None else 0
                        y = int(y) if y is not None else 0
                        cells[cell_name] = (int(x), int(y))
                        ###
                        if "SOURCE TIMING" in props:
                            source_timing_cells.add(cell_name)
            # ---------- NETS ----------
            elif state == 'NET':
                if line.startswith('-'):
                    # 進入新的一條 net
                    # e.g.  - net12 ( u1 CLK ) ( u2 A )
                    if current_net:
                        nets.append((current_net, current_pins))
                    current_net  = line.split()[1]
                    current_pins = []
                for cell, pin in re.findall(r'\(\s*([^\s()]+)\s+([^\s()]+)\s*\)', line):
                    current_pins.append((cell, pin))

    print(f"[INFO] 找到 {len(cells)} ({len(source_timing_cells)} + {len(cells)-len(source_timing_cells)}) 個單元 和 {len(nets)} 個網路\n")
    return cells, nets

def calculate_wirelength(cells, nets):
    """計算總線長(WL)，使用半周長(HPWL)方法"""
    print("開始計算線長(WL)...")
    total_wirelength = 0

    for net_name, pins in nets:
        if len(pins) <= 1:
            continue  # 忽略只有一個引腳的網路

        # 找出所有引腳的位置
        pin_positions = []
        for cell, pin in pins:
            if cell in cells:
                pin_positions.append(cells[cell])

        if not pin_positions:
            continue

        # 計算半周長
        min_x = min(pos[0] for pos in pin_positions)
        max_x = max(pos[0] for pos in pin_positions)
        min_y = min(pos[1] for pos in pin_positions)
        max_y = max(pos[1] for pos in pin_positions)

        wirelength = (max_x - min_x) + (max_y - min_y)
        total_wirelength += wirelength

    print(f"[INFO] 總線長: {total_wirelength}\n")
    return total_wirelength

def calculate_manhattan_distance(initial_cells, final_cells):
    """計算平均曼哈頓距離(D)"""
    total_distance = 0
    cell_count = 0

    print("[INFO] 計算平均曼哈頓距離(D)...")
    for cell_name, final_pos in final_cells.items():
        if cell_name in initial_cells:
            initial_pos = initial_cells[cell_name]
            # 計算曼哈頓距離
            distance = abs(final_pos[0] - initial_pos[0]) + abs(final_pos[1] - initial_pos[1])
            total_distance += distance
            cell_count += 1

    # 計算平均曼哈頓距離
    if cell_count > 0:
        avg_distance = total_distance / cell_count
    else:
        avg_distance = 0

    print(f"[INFO] 平均曼哈頓距離: {avg_distance:.6f} (基於 {cell_count} 個單元)\n")
    return avg_distance, cell_count

def normalize(metric_name, final, init):
    '''將各種度量標準正規化'''
    if init != 0:
        return 1 * (abs(init) - abs(final)) / abs(init)
    if metric_name in ("tns", "power", "wl"):
        return 0
    if metric_name == "d":
        return final / 1000
    if metric_name in "r":
        return final / 3600
    raise ValueError(f"Unknown metric: {metric_name}")

def calculate_score(norm, args):
    """計算最終得分 S = (1000 × P) − (50 × D) − (30 × R)"""
    p_value = args.alpha * norm['tns'] + args.beta * norm['power'] + args.gamma * norm['wl']
    score = 1000 * p_value - 50 * norm['d'] - 30 * norm['r']
    print(f"[INFO] P值: {p_value:.6f} ({args.alpha} * {norm['tns']:.6f} + {args.beta} * {norm['power']:.6f} + {args.gamma} * {norm['wl']:.6f})\n")
    print(f"[INFO] 最終得分: {score:.6f} (1000 * {p_value:.6f} - 50 * {norm['d']:.6f} - 30 * {norm['r']:.6f})\n")
    return score, p_value

def output_excel(team_name, design_name, metrics, args):
    '''將計算結果寫入 Excel 檔案'''
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Border, Side
    from openpyxl.utils import get_column_letter

    print(f"[INFO] 將結果寫入 Excel 檔案: {args.output_excel}")
    file_path = args.output_excel

    # 建立或讀取 Excel 檔案
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["team", "metric", "visible benchmarks"])
        ws.append(["", "", design_name])
    else:
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Results"]
        if design_name not in [ws.cell(row=2, column=i).value for i in range(3, ws.max_column + 1)]:
            ws.cell(row=2, column=ws.max_column + 1, value=design_name)

    # 找出該 design_name 要寫入的欄位編號
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=2, column=col).value == design_name:
            design_col = col
            break

    # 找出要從哪一列開始寫資料
    metric_to_row = {}
    for row in range(3, ws.max_row + 1):
        metric = ws.cell(row=row, column=2).value
        if metric:
            metric_to_row[metric] = row

    # 建立統一的格式
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_bottom = Border(
        bottom=Side(style='medium'),
        left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin')
    )
    light_blue = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    blue = PatternFill(start_color="66B2FF", end_color="66B2FF", fill_type="solid")
    light_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # 準備寫入資料
    written_metrics = set()
    for metric, value in metrics:
        # 數值處理：四捨五入
        if metric in ["TNS_norm", "Power_norm", "WL_norm", "D_norm(D)", "R_norm(R)", "P", "S"]:
            try:
                value = f"{float(value):.6f}"
            except:
                pass
        elif metric in ["total power (1e-2pW)", "Tpower_initial (1e-2pW)"]:
            try:
                value = f"{float(float(value)*100):.2f}"  # 從 1e-2pW 轉為 pW 單位顯示
            except:
                pass
        else:
            try:
                value = f"{float(value):.2f}"
            except:
                pass

        # 如果此 metric 已經存在就覆蓋，不存在就新增
        if metric in metric_to_row:
            row = metric_to_row[metric]
        else:
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=team_name)
            ws.cell(row=row, column=2, value=metric)
            metric_to_row[metric] = row

        # 寫入值
        ws.cell(row=row, column=design_col, value=value)

        written_metrics.add(metric)

    # 樣式處理：標題行
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = light_blue
            cell.font = Font(bold=True)
            cell.border = border

    # 樣式處理：第一欄 team 名稱欄上色（合併）
    ws.merge_cells(start_row=3, start_column=1, end_row=ws.max_row, end_column=1)
    cell = ws.cell(row=3, column=1)
    cell.value = team_name
    cell.fill = blue
    cell.font = Font(bold=True)
    cell.border = border

    # 加框線（整個表）
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # 加粗下框線
    for row_idx in [2, 8, 12, 18, 19, 21]:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col).border = thick_bottom

    # 顏色處理：initial 指標列（第 9–12 行）淺藍、19 行淺紅
    for r in range(9, 13):  # 9,10,11,12
        for c in range(2, ws.max_column + 1):  # 避開第1欄 team 的合併藍底
            ws.cell(row=r, column=c).fill = light_blue

    for c in range(2, ws.max_column + 1):
        ws.cell(row=19, column=c).fill = light_red

    # 自動調整欄寬
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)
    print(f"[INFO] Score for {design_name} 已寫入 {file_path}\n")

def main():
    parser = argparse.ArgumentParser(description='PPA評分計算器 - 計算最終得分(S)')
    parser.add_argument('--seed_def',                   required=True,  help='初始DEF檔案路徑')
    parser.add_argument('--sol_def',                    required=True,  help='優化後DEF檔案路徑')
    parser.add_argument('--alpha',          type=float, default= 0.34,  help='TNS權重')
    parser.add_argument('--beta',           type=float, default=-0.33,  help='功耗權重')
    parser.add_argument('--gamma',          type=float, default=-0.33,  help='線長權重')
    parser.add_argument('--runtime',        type=float, required=True,  help='運行時間比值')
    parser.add_argument('--start_time',     type=int,   default= 0,     help='開始時間')
    parser.add_argument('--report_script',              required=True,  help='OpenROAD報告腳本路徑')
    parser.add_argument('--output_excel',               required=True,  help='輸出EXCEL檔案路徑')
    parser.add_argument('--exit_code',      type=int,   default= 9,     help='退出代碼')
    args = parser.parse_args()

    # 其他參數
    team_name = "cadc1006"
    design_name = os.path.basename(args.seed_def).replace(".def", "")
    args.sdc_file = os.path.join(os.path.dirname(args.seed_def), f"{design_name}.sdc")
    d_baseline = 0.0
    r_baseline = 0.0

    # 執行 OpenROAD 並解析報告
    env_vars = {
        "SEED_DEF": args.seed_def,
        "SOL_DEF": args.sol_def,
        "SDC": args.sdc_file
    }
    report = get_openroad_report(args.report_script, env_vars)

    # 解析DEF檔案
    initial_cells, initial_nets = parse_def(args.seed_def)
    final_cells, final_nets     = parse_def(args.sol_def)

    # 計算初始和最終線長(WL)
    initial_wl  = calculate_wirelength(initial_cells, initial_nets)
    final_wl    = calculate_wirelength(final_cells, final_nets)

    # 計算平均曼哈頓距離(D)
    d_value, cell_count = calculate_manhattan_distance(initial_cells, final_cells)

    # 計算線長改進
    norm = {
        "tns": normalize("tns", report["tns_final"], report["tns_init"]),
        "power": normalize("power", report["power_final"], report["power_init"]),
        "wl": normalize("wl", final_wl, initial_wl),
        "d": normalize("d", d_value, d_baseline),
        "r": normalize("r", args.runtime, r_baseline)
    }
    print(f"[INFO] TNS改進(TNS_norm): {norm['tns']:.6f}")
    print(f"[INFO] 功耗改進(Power_norm): {norm['power']:.6f}")
    print(f"[INFO] 線長改進(WL_norm): {norm['wl']:.6f}")
    print(f"[INFO] 曼哈頓距離改進(D_norm): {norm['d']:.6f}")
    print(f"[INFO] 運行時間改進(R_norm): {norm['r']:.6f}")

    # 計算最終得分
    score, p_value = calculate_score(norm, args)

    # 輸出報告
    metrics = [
        ("Runtime of code (s)", args.runtime),
        ("TNS (ps)", report["tns_final"]),
        ("WNS (ps)", report["wns_final"]),
        ("total power (1e-2pW)", report["power_final"]),
        ("HPWL", final_wl),
        ("avg displacement", d_value),
        ("TNS_initial (ps)", report["tns_init"]),
        ("WNS_initial (ps)", report["wns_init"]),
        ("Tpower_initial (1e-2pW)", report["power_init"]),
        ("HPWL_initial", initial_wl),
        ("TNS_norm", norm['tns']),
        ("Power_norm", norm['power']),
        ("WL_norm", norm['wl']),
        ("D_norm(D)", norm['d']),
        ("R_norm(R)", norm['r']),
        ("P", p_value),
        ("S", score),
        ("START_TIME", datetime.datetime.fromtimestamp(args.start_time).strftime("%Y-%m-%d %H:%M:%S")),
        ("EXIT_CODE", args.exit_code),
        ("AREA_INIT", report["area_init"]),
        ("UTIL_INIT", report["util_init"]),
        ("AREA_FINAL", report["area_final"]),
        ("UTIL_FINAL", report["util_final"])
    ]

    output_excel(team_name, design_name, metrics, args)

if __name__ == "__main__":
    main()
